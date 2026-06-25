import platform
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from selenium import webdriver

REPORTS_DIR = Path(__file__).resolve().parents[1] / "reports"
REPORT_FILENAME = "dakota_chrome_extension_results.xlsx"
LATEST_SEARCH_REPORT_SENTINEL = REPORTS_DIR / ".latest_dakota_search_report.txt"

_session_report_path: Path | None = None

PERFORMANCE_REPORT_TEST_NAMES = frozenset({
    "test_dakota_search_time",
    "test_company_detail_loading_time",
    "test_company_contacts_loading_time",
    "test_company_type_specific_tab_loading_time",
    "test_search_load_more_time",
})

TIMING_NOT_APPLICABLE = "-"
ROW_TYPE_ITERATION = "Iteration"
ROW_TYPE_RUN_SUMMARY = "Run summary"
REPORT_FILE_SUFFIX = ".xlsx"
WORKSHEET_TITLE = "Results"

REPORT_COLUMNS = (
    "Row Type",
    "Tab",
    "Sample #",
    "Time (s)",
    "Min (s)",
    "Max (s)",
    "Performance Benchmark (s)",
    "Result",
    "Browser",
    "Recorded At",
    "Platform",
)

_RESULT_COLUMN_INDEX = REPORT_COLUMNS.index("Result") + 1
_ROW_TYPE_COLUMN_INDEX = REPORT_COLUMNS.index("Row Type") + 1
_TAB_COLUMN_INDEX = REPORT_COLUMNS.index("Tab") + 1

_NUMERIC_COLUMNS = frozenset({
    "Sample #",
    "Time (s)",
    "Min (s)",
    "Max (s)",
    "Performance Benchmark (s)",
})
_INTEGER_COLUMNS = frozenset({"Sample #"})
_DECIMAL_NUMBER_FORMAT = "0.000"

_HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E79")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_SUMMARY_ROW_FILL = PatternFill(fill_type="solid", fgColor="D6DCE4")
_ITERATION_FILL_A = PatternFill(fill_type="solid", fgColor="E2EFDA")
_ITERATION_FILL_B = PatternFill(fill_type="solid", fgColor="DDEBF7")
_RESULT_PASS_FILL = PatternFill(fill_type="solid", fgColor="C6EFCE")
_RESULT_FAIL_FILL = PatternFill(fill_type="solid", fgColor="FFC7CE")
_RESULT_NA_FILL = PatternFill(fill_type="solid", fgColor="FFEB9C")

_THIN_BORDER = Border(
    left=Side(style="thin", color="B4B4B4"),
    right=Side(style="thin", color="B4B4B4"),
    top=Side(style="thin", color="B4B4B4"),
    bottom=Side(style="thin", color="B4B4B4"),
)
_CENTER = Alignment(horizontal="center", vertical="center")
_LEFT = Alignment(horizontal="left", vertical="center")


def _coerce_cell_value(column_name: str, raw: str | None):
    if raw is None:
        return None
    text = str(raw).strip()
    if text == "":
        return None
    if column_name not in _NUMERIC_COLUMNS or text == TIMING_NOT_APPLICABLE:
        return text
    try:
        if column_name in _INTEGER_COLUMNS:
            return int(text)
        return float(text)
    except ValueError:
        return text


def _ensure_xlsx_report_path(report_path: Path) -> Path:
    if report_path.suffix.lower() == ".csv":
        return report_path.with_suffix(REPORT_FILE_SUFFIX)
    if report_path.suffix.lower() != REPORT_FILE_SUFFIX:
        return report_path.with_suffix(REPORT_FILE_SUFFIX)
    return report_path


def _is_run_summary_row(row_type: str) -> bool:
    return row_type.strip() == ROW_TYPE_RUN_SUMMARY


def _metadata_columns(metadata: "SearchReportMetadata") -> dict[str, str]:
    return {
        "Browser": metadata.browser,
        "Recorded At": metadata.recorded_at,
        "Platform": metadata.platform,
    }


def _iteration_row(
    tab: str,
    sample_number: int,
    elapsed_s: float | str,
    metadata: "SearchReportMetadata",
) -> dict[str, str]:
    time_value = (
        f"{elapsed_s:.3f}" if isinstance(elapsed_s, (int, float)) else str(elapsed_s)
    )
    return {
        "Row Type": ROW_TYPE_ITERATION,
        "Tab": tab,
        "Sample #": str(sample_number),
        "Time (s)": time_value,
        "Min (s)": "",
        "Max (s)": "",
        "Performance Benchmark (s)": "",
        "Result": "",
        **_metadata_columns(metadata),
    }


def _run_summary_row(
    tab: str,
    average_s: float | str,
    min_s: float | str,
    max_s: float | str,
    benchmark_s: float,
    result: str,
    metadata: "SearchReportMetadata",
) -> dict[str, str]:
    def _fmt(value: float | str) -> str:
        if isinstance(value, (int, float)):
            return f"{value:.3f}"
        return str(value)

    return {
        "Row Type": ROW_TYPE_RUN_SUMMARY,
        "Tab": tab,
        "Sample #": "",
        "Time (s)": _fmt(average_s),
        "Min (s)": _fmt(min_s),
        "Max (s)": _fmt(max_s),
        "Performance Benchmark (s)": f"{benchmark_s:.3f}",
        "Result": result,
        **_metadata_columns(metadata),
    }


def _summary_row_font(*, result: str | None = None) -> Font:
    if result is not None:
        normalized = result.strip().upper()
        if normalized == "PASS":
            return Font(bold=True, color="006100")
        if normalized == "FAIL":
            return Font(bold=True, color="9C0006")
        if normalized == "N/A":
            return Font(bold=True, color="9C6500")
    return Font(bold=True)


def _write_rows_to_worksheet(worksheet, rows: list[dict[str, str]], *, start_row: int) -> None:
    for offset, row in enumerate(rows):
        excel_row = start_row + offset
        for column_index, column_name in enumerate(REPORT_COLUMNS, start=1):
            raw = row.get(column_name, "")
            value = _coerce_cell_value(column_name, raw if raw != "" else None)
            cell = worksheet.cell(row=excel_row, column=column_index, value=value)
            if isinstance(value, float):
                cell.number_format = _DECIMAL_NUMBER_FORMAT


def _apply_report_formatting(worksheet) -> None:
    """Colourful header, alternating iteration rows, highlighted PASS/FAIL summaries."""
    max_col = len(REPORT_COLUMNS)
    numeric_col_indexes = {
        REPORT_COLUMNS.index(name) + 1
        for name in REPORT_COLUMNS
        if name in _NUMERIC_COLUMNS
    }

    for column_index in range(1, max_col + 1):
        header_cell = worksheet.cell(row=1, column=column_index)
        header_cell.fill = _HEADER_FILL
        header_cell.font = _HEADER_FONT
        header_cell.alignment = _CENTER
        header_cell.border = _THIN_BORDER

    last_tab = None
    use_fill_a = True

    for row_index in range(2, worksheet.max_row + 1):
        row_type = str(worksheet.cell(row=row_index, column=_ROW_TYPE_COLUMN_INDEX).value or "")
        is_summary = _is_run_summary_row(row_type)
        tab_value = str(worksheet.cell(row=row_index, column=_TAB_COLUMN_INDEX).value or "")

        if not is_summary and tab_value and tab_value != last_tab:
            if last_tab is not None:
                use_fill_a = not use_fill_a
            last_tab = tab_value
        elif is_summary:
            last_tab = None

        for column_index in range(1, max_col + 1):
            cell = worksheet.cell(row=row_index, column=column_index)
            cell.border = _THIN_BORDER

            if is_summary:
                cell.fill = _SUMMARY_ROW_FILL
                if column_index == _RESULT_COLUMN_INDEX:
                    result = str(cell.value or "")
                    cell.font = _summary_row_font(result=result)
                    if result.upper() == "PASS":
                        cell.fill = _RESULT_PASS_FILL
                    elif result.upper() == "FAIL":
                        cell.fill = _RESULT_FAIL_FILL
                    elif result.upper() == "N/A":
                        cell.fill = _RESULT_NA_FILL
                else:
                    cell.font = _summary_row_font()
                cell.alignment = _CENTER if column_index in numeric_col_indexes else _LEFT
            else:
                cell.fill = _ITERATION_FILL_A if use_fill_a else _ITERATION_FILL_B
                cell.font = Font(size=10)
                cell.alignment = _CENTER if column_index in numeric_col_indexes else _LEFT

    if worksheet.max_row >= 1:
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = f"A1:{get_column_letter(max_col)}{worksheet.max_row}"


def _autosize_worksheet_columns(worksheet) -> None:
    for column_index, column_name in enumerate(REPORT_COLUMNS, start=1):
        max_length = len(column_name)
        for row_index in range(1, worksheet.max_row + 1):
            value = worksheet.cell(row=row_index, column=column_index).value
            if value is not None:
                max_length = max(max_length, len(str(value)))
        worksheet.column_dimensions[get_column_letter(column_index)].width = min(max_length + 3, 42)


@dataclass(frozen=True)
class SearchSampleMetrics:
    elapsed_s: float
    api_response_time_s: float = 0.0


@dataclass(frozen=True)
class SearchReportMetadata:
    browser: str
    recorded_at: str
    platform: str


@dataclass(frozen=True)
class SearchSummary:
    tab: str
    average_time_s: float
    min_time_s: float
    max_time_s: float
    benchmark_s: float
    result: str
    average_api_response_time_s: float | None = None


def get_search_report_metadata(driver: webdriver.Chrome | None = None) -> SearchReportMetadata:
    browser_label = "Chrome"
    if driver is not None:
        try:
            version = driver.capabilities.get("browserVersion", "")
            browser_label = f"Chrome {version}".strip()
        except Exception:
            pass

    return SearchReportMetadata(
        browser=browser_label,
        recorded_at=datetime.now().strftime("%Y-%m-%d"),
        platform=platform.system().lower(),
    )


def build_search_report_rows(
    metrics_by_term: dict[str, list[SearchSampleMetrics]],
    benchmark_s: float,
    metadata: SearchReportMetadata,
) -> tuple[list[dict[str, str]], list[SearchSummary]]:
    rows: list[dict[str, str]] = []
    summaries: list[SearchSummary] = []

    for term, metrics in metrics_by_term.items():
        timings = [metric.elapsed_s for metric in metrics]
        for sample_number, metric in enumerate(metrics, start=1):
            rows.append(_iteration_row(term, sample_number, metric.elapsed_s, metadata))

        average_time_s = sum(timings) / len(timings)
        min_time_s = min(timings)
        max_time_s = max(timings)
        result = "PASS" if average_time_s <= benchmark_s else "FAIL"
        summaries.append(
            SearchSummary(
                tab=term,
                average_time_s=average_time_s,
                min_time_s=min_time_s,
                max_time_s=max_time_s,
                benchmark_s=benchmark_s,
                result=result,
            )
        )
        rows.append(
            _run_summary_row(term, average_time_s, min_time_s, max_time_s, benchmark_s, result, metadata)
        )

    return rows, summaries


def write_search_report_csv(report_path: Path, rows: list[dict[str, str]]) -> None:
    report_path = _ensure_xlsx_report_path(report_path)
    report_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = WORKSHEET_TITLE

    for column_index, column_name in enumerate(REPORT_COLUMNS, start=1):
        worksheet.cell(row=1, column=column_index, value=column_name)

    _write_rows_to_worksheet(worksheet, rows, start_row=2)
    _apply_report_formatting(worksheet)
    _autosize_worksheet_columns(worksheet)
    workbook.save(report_path)


def _new_performance_report_path() -> Path:
    return REPORTS_DIR / REPORT_FILENAME


def initialize_performance_report_for_session() -> Path:
    global _session_report_path
    report_path = _new_performance_report_path()
    write_search_report_csv(report_path, [])
    save_latest_search_report_path(report_path)
    _session_report_path = report_path
    return report_path


def get_performance_report_path() -> Path:
    global _session_report_path
    if _session_report_path is not None:
        path = _ensure_xlsx_report_path(_session_report_path)
        if path.is_file():
            return path
    return initialize_performance_report_for_session()


def save_latest_search_report_path(report_path: Path) -> None:
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    LATEST_SEARCH_REPORT_SENTINEL.write_text(str(report_path.resolve()), encoding="utf-8")


def read_latest_search_report_path() -> Path | None:
    if not LATEST_SEARCH_REPORT_SENTINEL.exists():
        return None
    text = LATEST_SEARCH_REPORT_SENTINEL.read_text(encoding="utf-8").strip()
    if not text:
        return None
    path = Path(text)
    if path.is_file():
        return path
    xlsx_path = _ensure_xlsx_report_path(path)
    return xlsx_path if xlsx_path.is_file() else None


def append_rows_to_search_report_csv(
    report_path: Path | None,
    rows: list[dict[str, str]],
) -> Path:
    if not rows:
        return get_performance_report_path()

    if report_path is None:
        report_path = get_performance_report_path()
    report_path = _ensure_xlsx_report_path(report_path)
    if not report_path.is_file():
        report_path = get_performance_report_path()
        report_path = _ensure_xlsx_report_path(report_path)

    workbook = load_workbook(report_path)
    worksheet = workbook.active
    _write_rows_to_worksheet(worksheet, rows, start_row=worksheet.max_row + 1)
    _apply_report_formatting(worksheet)
    _autosize_worksheet_columns(worksheet)
    workbook.save(report_path)
    return report_path


def build_company_detail_timing_row(
    search_term: str,
    sample_number: int,
    elapsed_s: float,
    metadata: SearchReportMetadata,
    **_kwargs,
) -> dict[str, str]:
    return _iteration_row(search_term, sample_number, elapsed_s, metadata)


def build_contacts_load_timing_row(
    search_term: str,
    sample_number: int,
    elapsed_s: float,
    metadata: SearchReportMetadata,
    **_kwargs,
) -> dict[str, str]:
    return _iteration_row(search_term, sample_number, elapsed_s, metadata)


def build_tab_load_timing_row(
    search_term: str,
    sample_number: int,
    elapsed_s: float,
    metadata: SearchReportMetadata,
    **_kwargs,
) -> dict[str, str]:
    return _iteration_row(search_term, sample_number, elapsed_s, metadata)


def build_search_load_more_skipped_timing_row(
    search_term: str,
    sample_number: int,
    metadata: SearchReportMetadata,
) -> dict[str, str]:
    return _iteration_row(search_term, sample_number, TIMING_NOT_APPLICABLE, metadata)


def build_search_load_more_timing_row(
    search_term: str,
    sample_number: int,
    elapsed_s: float,
    metadata: SearchReportMetadata,
    **_kwargs,
) -> dict[str, str]:
    return _iteration_row(search_term, sample_number, elapsed_s, metadata)


def _summary_from_timings(
    tab: str,
    timings: list[float],
    benchmark_s: float,
    metadata: SearchReportMetadata,
) -> tuple[dict[str, str], SearchSummary]:
    average_time_s = sum(timings) / len(timings)
    min_time_s = min(timings)
    max_time_s = max(timings)
    result = "PASS" if average_time_s <= benchmark_s else "FAIL"
    summary = SearchSummary(
        tab=tab,
        average_time_s=average_time_s,
        min_time_s=min_time_s,
        max_time_s=max_time_s,
        benchmark_s=benchmark_s,
        result=result,
    )
    row = _run_summary_row(
        tab, average_time_s, min_time_s, max_time_s, benchmark_s, result, metadata
    )
    return row, summary


def build_search_load_more_skipped_run_summary_for_term(
    term: str,
    benchmark_s: float,
    metadata: SearchReportMetadata,
) -> tuple[dict[str, str], SearchSummary]:
    summary = SearchSummary(
        tab=term,
        average_time_s=0.0,
        min_time_s=0.0,
        max_time_s=0.0,
        benchmark_s=benchmark_s,
        result="N/A",
    )
    row = _run_summary_row(
        term, TIMING_NOT_APPLICABLE, TIMING_NOT_APPLICABLE, TIMING_NOT_APPLICABLE,
        benchmark_s, "N/A", metadata,
    )
    return row, summary


def build_search_load_more_run_summary_for_term(
    term: str,
    timings: list[float],
    benchmark_s: float,
    metadata: SearchReportMetadata,
    **_,
) -> tuple[dict[str, str], SearchSummary]:
    if not timings:
        raise ValueError(f"No timings recorded for term {term!r}")
    return _summary_from_timings(term, timings, benchmark_s, metadata)


def build_company_detail_run_summary_for_term(
    term: str,
    timings: list[float],
    benchmark_s: float,
    metadata: SearchReportMetadata,
    *_,
    **__,
) -> tuple[dict[str, str], SearchSummary]:
    if not timings:
        raise ValueError(f"No timings recorded for term {term!r}")
    return _summary_from_timings(term, timings, benchmark_s, metadata)


def build_company_detail_run_summary_rows(
    benchmark_terms: tuple[str, ...],
    detail_elapsed_by_term: dict[str, list[float]],
    detail_view_by_term: dict[str, tuple[str, str]],
    benchmark_s: float,
    metadata: SearchReportMetadata,
    **_,
) -> tuple[list[dict[str, str]], list[SearchSummary]]:
    rows: list[dict[str, str]] = []
    summaries: list[SearchSummary] = []

    for term in benchmark_terms:
        timings = detail_elapsed_by_term.get(term) or []
        if not timings:
            continue
        row, summary = build_company_detail_run_summary_for_term(
            term, timings, benchmark_s, metadata
        )
        rows.append(row)
        summaries.append(summary)

    return rows, summaries
